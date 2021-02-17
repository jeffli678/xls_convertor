#!/usr/bin/python3
from openpyxl import load_workbook, Workbook
import shutil
from config import *
from copy import copy
import re

def calc_ylzj(s):

    expr = r'\s(\d+)/(C|T|X)M'
    m = re.search(expr, s)
    if m:
        return m.group(1)

    expr = r'TDA-(\d+)'
    m = re.search(expr, s)
    if m:
        return m.group(1)

    # if 'CCD' in s or 'CCK' in s:
    if s.find('7/7') >= 0:
        return '180'
    if s.find('8/8') >= 0:
        return '200'
    if s.find('9/7') >= 0:
        return '252'
    if s.find('9/9') >= 0:
        return '252'
    if s.find('10/8') >= 0:
        return '282'
    if s.find('10/10') >= 0:
        return '282'
    if s.find('12/9') >= 0:
        return '317'
    if s.find('12/12') >= 0:
        return '317'
    if s.find('15/11') >= 0:
        return '383'
    if s.find('15/15') >= 0:
        return '383'
    if s.find('18/13') >= 0:
        return '450'
    if s.find('18/18') >= 0:
        return '450'

    return ''

def main():
    wb = load_workbook(filename = input_file, data_only = True)
    sh = wb['Sheet1']

    shutil.copyfile(template_file, output_file)
    output_wb = load_workbook(filename = output_file)
    output_ws = output_wb['Sheet1']

    template_wb = load_workbook(filename = template_file)
    template_ws = template_wb['Sheet1']

    for idx in range(start_row, end_row + 1):
        for row in range(template_start_row, template_end_row + 1):
            for col in range(1, template_columns + 1):
                output_value = ''
                template_cell = template_ws.cell(row = row, column = col)
                template_value = template_cell.value
                if type(template_value) == str and template_value.startswith('{{YLZJ(') and template_value.endswith(')}}'):
                    column_str = template_value.replace('{{YLZJ(', '').replace(')}}', '')
                    ylzj_input = sh[column_str + str(idx)].value
                    output_value = calc_ylzj(ylzj_input)

                elif type(template_value) == str and template_value.startswith('{{') and template_value.endswith('}}'):
                    column_str = template_value.replace('{{', '').replace('}}', '')
                    output_value = sh[column_str + str(idx)].value
                else:
                    output_value = template_value
                
                output_row = (idx - start_row) * (template_end_row - template_start_row + 1) + row
                output_cell = output_ws.cell(row = output_row, column = col, value = output_value)
                if template_cell.has_style:
                    output_cell.font = copy(template_cell.font)
                    output_cell.border = copy(template_cell.border)
                    output_cell.fill = copy(template_cell.fill)
                    output_cell.number_format = copy(template_cell.number_format)
                    output_cell.protection = copy(template_cell.protection)
                    output_cell.alignment = copy(template_cell.alignment)

                output_ws.row_dimensions[output_row].height = template_ws.row_dimensions[row].height
                
    output_wb.save(filename = output_file)


if __name__ == '__main__':
    main()